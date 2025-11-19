import pandas as pd
import openpyxl
import os
from openpyxl.styles import Font
from openpyxl import Workbook
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image

def create_data_audit(df: pd.DataFrame, wb: Workbook) -> Workbook:
    audit = df.describe().T

    n = len(df)
    audit['NUM_MISSING'] = df.isna().sum()
    audit['%_MISSING'] = (audit['NUM_MISSING'] / n).round(2)
    audit['NUM_UNIQUE'] = df.nunique(dropna=True)
    audit['%_UNIQUE'] = (audit['NUM_UNIQUE'] / n).round(2)
    audit['Comment'] = '… TODO: fill in with analysis …'

    # A strange 'Unnamed: 0' row appears in the resulting table
    # From where does it come?
    audit = audit[audit.index != 'Unnamed: 0']

    audit.columns = [str(c).upper() for c in audit.columns]

    if "data_audit" in wb.sheetnames:
        del wb["data_audit"]

    ws = wb.create_sheet("data_audit")
    header_font = Font(bold=True, sz=13)
    content_font = Font(sz=12)
    ws.cell(row=1, column=1, value="COLUMN").font = header_font

    for col_idx, stat_name in enumerate(audit.columns, start=2):
        cell = ws.cell(row=1, column=col_idx,
                       value=stat_name).font = header_font

    for row_idx, (feature_name, stats_row) in enumerate(audit.iterrows(),
                                                        start=2):

        ws.cell(row=row_idx, column=1, value=feature_name).font = content_font

        for col_idx, stat_value in enumerate(stats_row, start=2):
            try:
                ws.cell(row=row_idx, column=col_idx,
                        value=float(stat_value)).font = content_font
            except:
                ws.cell(row=row_idx, column=col_idx,
                        value=stat_value).font = content_font

    return wb

def create_features_analysis(df: pd.DataFrame, wb: Workbook) -> Workbook:
    header_font = Font(bold=True, sz=13)
    content_font = Font(sz=12)

    for col in df.columns:
        if col == 'Unnamed: 0':
            continue

        series = df[col]

        sheet_name = str(col)
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)

        value_counts = series.value_counts(dropna=False)
        value_counts = value_counts.sort_index()

        ws.cell(row=1, column=1, value=f'Column: {col}').font = header_font
        ws.cell(row=2, column=1, value="Value").font = header_font
        ws.cell(row=2, column=2, value='Count').font = header_font
        
        for row_idx, (val, cnt) in enumerate(value_counts.items(), start=3):
            ws.cell(row=row_idx, column=1, value=str(val)).font = content_font
            ws.cell(row=row_idx, column=2, value=int(cnt)).font = content_font
        
        
        num_unique = value_counts.shape[0]

        fig, ax = plt.subplots(figsize=(6, 4))

        if num_unique > 10:
            ax.hist(series.to_numpy(), bins=30)
            ax.set_title(f"Histogram of {col}")
            ax.set_xlabel(col)
            ax.set_ylabel("Frequency")
        else:
            vc_for_plot = value_counts

            x_labels = [str(x) if not pd.isna(x) else "NaN"
                        for x in vc_for_plot.index]
            ax.bar(x_labels, vc_for_plot.to_numpy())
            ax.set_title(f"Bar chart of {col}")
            ax.set_xlabel(col)
            ax.set_ylabel("Count")
            ax.tick_params(axis='x', rotation=45)

        fig.tight_layout()
        
        img_path = f"plot_{sheet_name}.png"
        fig.savefig(os.path.join('tmp_plots/', img_path), dpi=120)
        plt.close(fig)

        img = Image(os.path.join('tmp_plots/', img_path))
        ws.add_image(img, "E2")
    return wb


def main() -> None:
    output_file_name = 'data_audit.xlsx'
    output_file_path = 'Week_02 - Hello, Machine Learning/Data Science'
    output_path = os.path.join(output_file_path, output_file_name)
    df = pd.read_csv('DATA/telecom_churn_clean.csv')

    wb = openpyxl.load_workbook(output_path)

    wb = create_data_audit(df, wb)
    wb = create_features_analysis(df, wb)

    wb.save(output_path)



if __name__ == '__main__':
    main()
