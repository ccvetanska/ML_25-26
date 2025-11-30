import pandas as pd
import openpyxl
import os
from openpyxl.styles import Font
from openpyxl import Workbook
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from sklearn.neighbors import KNeighborsClassifier
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler


def create_neighboors_comparision_ws(X_train, X_test, y_train, y_test) -> None:
    output_file_name = 'model_report.xlsx'
    output_file_path = 'Week_02 - Hello, Machine Learning/Data Science'
    output_path = os.path.join(output_file_path, output_file_name)
    model_report_wb = openpyxl.load_workbook(output_path)

    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_test_scaled = scaler.transform(X_test)

    neighbors = range(1, 25)

    train_accuracies = {}
    test_accuracies = {}

    for k in neighbors:
        knn = KNeighborsClassifier(n_neighbors=k)
        knn.fit(X_train_scaled, y_train)

        train_accuracies[k] = knn.score(X_train_scaled, y_train)
        test_accuracies[k] = knn.score(X_test_scaled, y_test)

    plt.figure(figsize=(8, 6))
    plt.title('KNN models')

    plt.plot(list(neighbors),
             list(train_accuracies.values()),
             label='Training Accuracy')
    plt.plot(list(neighbors),
             list(test_accuracies.values()),
             label='Testing Accuracy')

    plt.legend()
    plt.xlabel('Number of Neighbors (k)')
    plt.ylabel('Accuracy')
    plt.grid(True)
    plt.savefig(
        'Week_02 - Hello, Machine Learning/Data Science/KNN_models.png')

    sheet_name = 'KNN_neighboors_comparision'

    if sheet_name in model_report_wb.sheetnames:
        ws_old = model_report_wb[sheet_name]
        model_report_wb.remove(ws_old)

    ws = model_report_wb.create_sheet(title=sheet_name)

    img = Image(
        'Week_02 - Hello, Machine Learning/Data Science/KNN_models.png')
    ws.add_image(img, "E1")

    ws['A1'] = 'k'
    ws['B1'] = 'train_accuracy'
    ws['C1'] = 'test_accuracy'

    row = 2
    for k in neighbors:
        ws[f'A{row}'] = k
        ws[f'B{row}'] = train_accuracies[k]
        ws[f'C{row}'] = test_accuracies[k]
        row += 1

    # Here I wanted to bold the best test accuracy but
    # instead the last row was bolded. Why?

    # best_k, best_score = max(test_accuracies.items())
    # best_row = list(neighbors).index(best_k) + 2
    # bold_font = Font(bold=True)
    # ws[f'A{best_row}'].font=bold_font
    # ws[f'B{best_row}'].font=bold_font
    # ws[f'C{best_row}'].font=bold_font

    model_report_wb.save(output_path)


def create_hyperparameter_tuning_ws(X_train, X_test, y_train, y_test) -> None:
    output_file_name = 'model_report.xlsx'
    output_file_path = 'Week_02 - Hello, Machine Learning/Data Science'
    output_path = os.path.join(output_file_path, output_file_name)
    model_report_wb = openpyxl.load_workbook(output_path)

    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_test_scaled = scaler.transform(X_test)

    neighbors = range(1, 10)
    leaf_size_vals = range(28, 32)
    p_vals = {1, 2, 3}

    results = []
    for k in neighbors:
        for leaf_size in leaf_size_vals:
            for p in p_vals:
                knn = KNeighborsClassifier(n_neighbors=k,
                                           leaf_size=leaf_size,
                                           p=p)
                knn.fit(X_train_scaled, y_train)
                train_score = knn.score(X_train_scaled, y_train)
                test_score = knn.score(X_test_scaled, y_test)
                results.append((k, leaf_size, p, train_score, test_score))

    sheet_name = 'KNN_hyperparameter_tuning'

    if sheet_name in model_report_wb.sheetnames:
        ws_old = model_report_wb[sheet_name]
        model_report_wb.remove(ws_old)

    ws = model_report_wb.create_sheet(title=sheet_name)

    ws['A1'] = 'k'
    ws['B1'] = 'leaf_size'
    ws['C1'] = 'p'
    ws['D1'] = 'train_accuracy'
    ws['E1'] = 'test_accuracy'

    row = 2
    for k, leaf_size, p, train_score, test_score in results:
        ws[f'A{row}'] = k
        ws[f'B{row}'] = leaf_size
        ws[f'C{row}'] = p
        ws[f'D{row}'] = train_score
        ws[f'E{row}'] = test_score
        row += 1

    model_report_wb.save(output_path)


def main():
    df = pd.read_csv('DATA/telecom_churn_clean.csv')
    X = df.drop("churn", axis=1)
    y = df["churn"]

    X_train, X_test, y_train, y_test = train_test_split(X,
                                                        y,
                                                        test_size=0.3,
                                                        random_state=21,
                                                        stratify=y)
    #create_neighboors_comparision_ws(X_train, X_test, y_train, y_test)
    create_hyperparameter_tuning_ws(X_train, X_test, y_train, y_test)


if __name__ == '__main__':
    main()
