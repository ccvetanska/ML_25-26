import pandas as pd
import openpyxl
import os
import sys 
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..')))

from openpyxl.styles import Font
from openpyxl import Workbook
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
from ml_lib.neighbors import KNeighborsClassifier
from ml_lib.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler


def create_neighboors_comparision_ws(X_train, X_test, y_train, y_test) -> None:
    output_file_name = 'ml_lib_model_report.xlsx'
    output_file_path = 'Week_02 - Hello, Machine Learning/Engineering'
    output_path = os.path.join(output_file_path, output_file_name)
    model_report_wb = openpyxl.load_workbook(output_path)

    scaler = StandardScaler()
    X_train_scaled_arr = scaler.fit_transform(X_train)
    X_test_scaled_arr = scaler.transform(X_test)
    X_train_scaled = pd.DataFrame(X_train_scaled_arr, columns=X_train.columns, index=X_train.index)
    X_test_scaled = pd.DataFrame(X_test_scaled_arr, columns=X_test.columns, index=X_test.index)

    neighbors = range(1, 25)

    train_accuracies = {}
    test_accuracies = {}

    for k in neighbors:
        knn = KNeighborsClassifier(n_neighbors=k)
        knn.fit(X_train_scaled, y_train)

        train_accuracies[k] = knn.score(X_train_scaled, y_train)
        test_accuracies[k] = knn.score(X_test_scaled, y_test)

    plt.figure(figsize=(8, 6))
    plt.title('KNN models (ml_lib)')

    plt.plot(list(neighbors),
             list(train_accuracies.values()),
             label='Training Accuracy')
    plt.plot(list(neighbors),
             list(test_accuracies.values()),
             label='Testing Accuracy')

    plt.legend()
    plt.xlabel('Number of Neighbors (k)')
    plt.ylabel('Accuracy')
    plt.grid(True)
    plt.savefig(
        'Week_02 - Hello, Machine Learning/Engineering/KNN_models_ml_lib.png')

    sheet_name = 'KNN_neighboors_comparision'

    if sheet_name in model_report_wb.sheetnames:
        ws_old = model_report_wb[sheet_name]
        model_report_wb.remove(ws_old)

    ws = model_report_wb.create_sheet(title=sheet_name)

    img = Image(
        'Week_02 - Hello, Machine Learning/Engineering/KNN_models_ml_lib.png')
    ws.add_image(img, "E1")

    ws['A1'] = 'k'
    ws['B1'] = 'train_accuracy'
    ws['C1'] = 'test_accuracy'

    row = 2
    for k in neighbors:
        ws[f'A{row}'] = k
        ws[f'B{row}'] = train_accuracies[k]
        ws[f'C{row}'] = test_accuracies[k]
        row += 1

    model_report_wb.save(output_path)

def main():
    df = pd.read_csv('DATA/telecom_churn_clean.csv')
    X = df.drop("churn", axis=1)
    y = df["churn"]

    X_train, X_test, y_train, y_test = train_test_split(X,
                                                        y,
                                                        test_size=0.3,
                                                        random_state=21,
                                                        stratify=y)
    create_neighboors_comparision_ws(X_train, X_test, y_train, y_test)


if __name__ == '__main__':
    main()
